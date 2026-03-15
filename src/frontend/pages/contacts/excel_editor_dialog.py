# FILE: src/frontend/pages/contacts/excel_editor_dialog.py
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from PySide6.QtCore import QAbstractTableModel, QModelIndex, Qt, QSortFilterProxyModel, QItemSelectionModel
from PySide6.QtGui import QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QTableView,
    QVBoxLayout,
    QScrollArea,
    QSizePolicy,
)
from openpyxl.utils import get_column_letter

from backend.integrations.excel.workbook_editor_io import (
    SheetGrid,
    WorkbookGrid,
    save_workbook_grid_to_xlsx,
    suggest_value_only_save_path,
)
from backend.integrations.windows.win_file_picker import Filter, pick_save_file


@dataclass
class CellEditCommand:
    sheet_index: int
    row: int
    col: int
    old_value: str
    new_value: str


@dataclass
class SheetSnapshotCommand:
    sheet_index: int
    before_rows: list[list[str]]
    after_rows: list[list[str]]
    label: str


EditCommand = CellEditCommand | SheetSnapshotCommand

EXCEL_SAMPLE_HEADERS = ["사번", "이름", "전화번호", "대리점명", "지사명"]


class SheetGridTableModel(QAbstractTableModel):
    def __init__(self, sheet: SheetGrid, on_cell_edited: Callable[[int, int, str, str], None] | None = None) -> None:
        super().__init__()
        self._sheet = sheet
        self._sheet.ensure_rectangular()
        self._on_cell_edited = on_cell_edited
        self._history_suspended = False

    @property
    def sheet(self) -> SheetGrid:
        return self._sheet

    def set_sheet(self, sheet: SheetGrid) -> None:
        self.beginResetModel()
        self._sheet = sheet
        self._sheet.ensure_rectangular()
        self.endResetModel()

    def set_history_suspended(self, suspended: bool) -> None:
        self._history_suspended = suspended

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return max(1, self._sheet.row_count)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        if parent.isValid():
            return 0
        return max(1, self._sheet.col_count)

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole):
        if not index.isValid():
            return None
        if role not in (Qt.DisplayRole, Qt.EditRole):
            return None
        row = index.row()
        col = index.column()
        if row >= len(self._sheet.rows) or col >= len(self._sheet.rows[row]):
            return ""
        return self._sheet.rows[row][col]

    def setData(self, index: QModelIndex, value, role: int = Qt.EditRole) -> bool:
        if role != Qt.EditRole or not index.isValid():
            return False
        self._ensure_cell(index.row(), index.column())
        old_value = self._sheet.rows[index.row()][index.column()]
        new_value = "" if value is None else str(value)
        if old_value == new_value:
            return True
        self._sheet.rows[index.row()][index.column()] = new_value
        self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
        if self._on_cell_edited and not self._history_suspended:
            self._on_cell_edited(index.row(), index.column(), old_value, new_value)
        return True

    def flags(self, index: QModelIndex):
        if not index.isValid():
            return Qt.ItemIsEnabled
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable

    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return get_column_letter(section + 1)
        return str(section + 1)

    def delete_rows(self, row_indexes: list[int]) -> int:
        rows = sorted({r for r in row_indexes if r >= 0}, reverse=True)
        if not rows:
            return 0
        self.beginResetModel()
        for row in rows:
            if row < len(self._sheet.rows):
                del self._sheet.rows[row]
        if not self._sheet.rows:
            self._sheet.rows = [[""]]
        self._sheet.ensure_rectangular()
        self.endResetModel()
        return len(rows)

    def delete_columns(self, col_indexes: list[int]) -> int:
        cols = sorted({c for c in col_indexes if c >= 0}, reverse=True)
        if not cols:
            return 0
        self.beginResetModel()
        self._sheet.ensure_rectangular()
        for row in self._sheet.rows:
            for col in cols:
                if col < len(row):
                    del row[col]
            if not row:
                row.append("")
        self._sheet.ensure_rectangular()
        self.endResetModel()
        return len(cols)

    def move_column(self, source_col: int, target_col: int) -> bool:
        self._sheet.ensure_rectangular()
        col_count = self.columnCount()
        if col_count <= 1:
            return False
        if source_col < 0 or source_col >= col_count:
            return False
        if target_col < 0 or target_col >= col_count:
            return False
        if source_col == target_col:
            return False

        self.beginResetModel()
        for row in self._sheet.rows:
            value = row.pop(source_col)
            row.insert(target_col, value)
        self._sheet.ensure_rectangular()
        self.endResetModel()
        return True

    def replace_rows(self, rows: list[list[str]]) -> None:
        self.beginResetModel()
        copied = [["" if v is None else str(v) for v in row] for row in rows]
        self._sheet.rows = copied if copied else [[""]]
        self._sheet.ensure_rectangular()
        self.endResetModel()

    def get_rows_copy(self) -> list[list[str]]:
        self._sheet.ensure_rectangular()
        return [list(row) for row in self._sheet.rows]

    def get_unique_values_with_count(self, col: int) -> list[tuple[str, int]]:
        self._sheet.ensure_rectangular()
        counts: dict[str, int] = {}
        for row in self._sheet.rows:
            value = row[col] if col < len(row) else ""
            counts[value] = counts.get(value, 0) + 1
        items = list(counts.items())
        items.sort(key=lambda x: (x[0] != "", x[0].casefold()))
        return items

    def find_next(self, needle: str, start_after: tuple[int, int] | None, case_sensitive: bool) -> tuple[int, int] | None:
        if needle == "":
            return None

        total_rows = self.rowCount()
        total_cols = self.columnCount()
        start_index = 0
        if start_after is not None:
            start_index = (start_after[0] * total_cols + start_after[1] + 1) % (total_rows * total_cols)

        target = needle if case_sensitive else needle.casefold()
        total_cells = total_rows * total_cols
        for offset in range(total_cells):
            linear = (start_index + offset) % total_cells
            row = linear // total_cols
            col = linear % total_cols
            value = self.index(row, col).data(Qt.DisplayRole) or ""
            source = value if case_sensitive else str(value).casefold()
            if target in source:
                return row, col
        return None

    def replace_all(self, needle: str, replacement: str, case_sensitive: bool) -> int:
        if needle == "":
            return 0

        self.beginResetModel()
        changed = 0
        self._sheet.ensure_rectangular()
        if case_sensitive:
            for r, row in enumerate(self._sheet.rows):
                for c, value in enumerate(row):
                    if needle in value:
                        new_value = value.replace(needle, replacement)
                        if new_value != value:
                            self._sheet.rows[r][c] = new_value
                            changed += 1
        else:
            needle_fold = needle.casefold()
            needle_len = len(needle)
            for r, row in enumerate(self._sheet.rows):
                for c, value in enumerate(row):
                    if needle_fold not in value.casefold():
                        continue
                    out: list[str] = []
                    i = 0
                    text = value
                    text_fold = text.casefold()
                    local_changed = False
                    while i < len(text):
                        if text_fold[i:i + needle_len] == needle_fold:
                            out.append(replacement)
                            i += needle_len
                            local_changed = True
                        else:
                            out.append(text[i])
                            i += 1
                    if local_changed:
                        self._sheet.rows[r][c] = "".join(out)
                        changed += 1
        self.endResetModel()
        return changed

    def add_combined_column(self, source_cols: list[int], header_title: str, separator: str = " ") -> int:
        self.beginResetModel()
        self._sheet.ensure_rectangular()
        col_index = self._sheet.col_count
        for row_idx, row in enumerate(self._sheet.rows):
            if row_idx == 0:
                row.append(header_title)
                continue
            parts: list[str] = []
            for col in source_cols:
                value = row[col] if col < len(row) else ""
                value = str(value).strip()
                if value:
                    parts.append(value)
            row.append(separator.join(parts))
        self._sheet.ensure_rectangular()
        self.endResetModel()
        return col_index

    def _ensure_cell(self, row: int, col: int) -> None:
        while len(self._sheet.rows) <= row:
            self._sheet.rows.append([])
        max_width = max(self._sheet.col_count, col + 1, 1)
        for r in self._sheet.rows:
            if len(r) < max_width:
                r.extend([""] * (max_width - len(r)))


class ColumnValueFilterProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._allowed_values_by_col: dict[int, set[str]] = {}

    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        model = self.sourceModel()
        if model is None:
            return True
        for col, allowed_values in self._allowed_values_by_col.items():
            if not allowed_values:
                return False
            idx = model.index(source_row, col, source_parent)
            value = idx.data(Qt.DisplayRole) or ""
            if str(value) not in allowed_values:
                return False
        return True

    def active_filters(self) -> dict[int, set[str]]:
        return {col: set(values) for col, values in self._allowed_values_by_col.items()}

    def set_allowed_values(self, column: int, allowed_values: set[str] | None) -> None:
        if allowed_values is None:
            self._allowed_values_by_col.pop(column, None)
        else:
            self._allowed_values_by_col[column] = {str(v) for v in allowed_values}
        self.invalidateFilter()

    def clear_all_filters(self) -> None:
        self._allowed_values_by_col.clear()
        self.invalidateFilter()

    def has_active_filters(self) -> bool:
        return len(self._allowed_values_by_col) > 0

    def prune_invalid_columns(self, column_count: int) -> None:
        self._allowed_values_by_col = {
            col: values
            for col, values in self._allowed_values_by_col.items()
            if col < column_count
        }
        self.invalidateFilter()

    def shift_after_deleted_columns(self, deleted_cols: list[int]) -> None:
        deleted = sorted({c for c in deleted_cols if c >= 0})
        if not deleted:
            return
        shifted: dict[int, set[str]] = {}
        for col, values in self._allowed_values_by_col.items():
            if col in deleted:
                continue
            shift = sum(1 for d in deleted if d < col)
            shifted[col - shift] = values
        self._allowed_values_by_col = shifted
        self.invalidateFilter()


class ColumnFilterDialog(QDialog):
    def __init__(
        self,
        model: SheetGridTableModel,
        active_filters: dict[int, set[str]],
        initial_column: int = 0,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._model = model
        self._active_filters = {col: set(values) for col, values in active_filters.items()}
        self._current_all_values: list[str] = []

        self.setWindowTitle("열 필터")
        self.resize(520, 620)

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        title = QLabel("열별 값 필터")
        title.setStyleSheet("font-size:15px; font-weight:700;")
        root.addWidget(title)

        desc = QLabel("열을 선택하면 해당 열에 존재하는 값 목록을 확인하고 체크로 포함/제외할 수 있습니다.")
        desc.setStyleSheet("color:#6b7280;")
        desc.setWordWrap(True)
        root.addWidget(desc)

        top = QHBoxLayout()
        top.addWidget(QLabel("대상 열"))
        self.column_combo = QComboBox()
        for col in range(max(1, self._model.columnCount())):
            self.column_combo.addItem(f"{get_column_letter(col + 1)}열", col)
        top.addWidget(self.column_combo, 1)
        root.addLayout(top)

        search_row = QHBoxLayout()
        self.value_search = QLineEdit()
        self.value_search.setPlaceholderText("값 검색")
        self.btn_select_all = QPushButton("전체 선택")
        self.btn_unselect_all = QPushButton("전체 해제")
        self.btn_clear_this = QPushButton("이 열 필터 해제")
        search_row.addWidget(self.value_search, 1)
        search_row.addWidget(self.btn_select_all)
        search_row.addWidget(self.btn_unselect_all)
        search_row.addWidget(self.btn_clear_this)
        root.addLayout(search_row)

        self.summary_label = QLabel("")
        self.summary_label.setStyleSheet("color:#6b7280;")
        root.addWidget(self.summary_label)

        self.value_list = QListWidget()
        root.addWidget(self.value_list, 1)

        bottom = QHBoxLayout()
        self.btn_cancel = QPushButton("취소")
        self.btn_apply = QPushButton("적용")
        bottom.addStretch(1)
        bottom.addWidget(self.btn_cancel)
        bottom.addWidget(self.btn_apply)
        root.addLayout(bottom)

        self.column_combo.currentIndexChanged.connect(self._reload_value_list)
        self.value_search.textChanged.connect(self._apply_search_filter)
        self.btn_select_all.clicked.connect(self._select_all_visible)
        self.btn_unselect_all.clicked.connect(self._unselect_all_visible)
        self.btn_clear_this.clicked.connect(self._clear_current_column_filter)
        self.btn_cancel.clicked.connect(self.reject)
        self.btn_apply.clicked.connect(self._accept_and_store)

        initial_index = max(0, min(initial_column, self.column_combo.count() - 1))
        self.column_combo.setCurrentIndex(initial_index)
        self._reload_value_list()

    def _current_column(self) -> int:
        data = self.column_combo.currentData()
        return int(data) if data is not None else 0

    def _reload_value_list(self) -> None:
        self.value_list.clear()
        self._current_all_values = []
        col = self._current_column()
        values_with_count = self._model.get_unique_values_with_count(col)
        self._current_all_values = [value for value, _count in values_with_count]

        active = self._active_filters.get(col)
        for value, count in values_with_count:
            label = "(빈값)" if value == "" else value
            item = QListWidgetItem(f"{label}  ({count})")
            item.setData(Qt.UserRole, value)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            checked = Qt.Checked if (active is None or value in active) else Qt.Unchecked
            item.setCheckState(checked)
            self.value_list.addItem(item)

        self._apply_search_filter(self.value_search.text())
        self._update_summary()

    def _apply_search_filter(self, text: str) -> None:
        query = (text or "").strip().casefold()
        for i in range(self.value_list.count()):
            item = self.value_list.item(i)
            raw = str(item.data(Qt.UserRole) or "")
            label = "(빈값)" if raw == "" else raw
            visible = (query == "") or (query in label.casefold())
            item.setHidden(not visible)
        self._update_summary()

    def _select_all_visible(self) -> None:
        for i in range(self.value_list.count()):
            item = self.value_list.item(i)
            if not item.isHidden():
                item.setCheckState(Qt.Checked)
        self._update_summary()

    def _unselect_all_visible(self) -> None:
        for i in range(self.value_list.count()):
            item = self.value_list.item(i)
            if not item.isHidden():
                item.setCheckState(Qt.Unchecked)
        self._update_summary()

    def _clear_current_column_filter(self) -> None:
        self._active_filters.pop(self._current_column(), None)
        for i in range(self.value_list.count()):
            item = self.value_list.item(i)
            item.setCheckState(Qt.Checked)
        self._update_summary()

    def _checked_values(self) -> set[str]:
        selected: set[str] = set()
        for i in range(self.value_list.count()):
            item = self.value_list.item(i)
            if item.checkState() == Qt.Checked:
                selected.add(str(item.data(Qt.UserRole) or ""))
        return selected

    def _update_summary(self) -> None:
        total = self.value_list.count()
        visible = sum(1 for i in range(total) if not self.value_list.item(i).isHidden())
        checked = sum(1 for i in range(total) if self.value_list.item(i).checkState() == Qt.Checked)
        self.summary_label.setText(f"전체 값 {total}개 / 표시 중 {visible}개 / 선택 {checked}개")

    def _accept_and_store(self) -> None:
        col = self._current_column()
        checked = self._checked_values()
        all_values = set(self._current_all_values)
        if checked == all_values:
            self._active_filters.pop(col, None)
        else:
            self._active_filters[col] = checked
        self.accept()

    def get_filters(self) -> dict[int, set[str]]:
        return {col: set(values) for col, values in self._active_filters.items()}


class CombinedColumnDialog(QDialog):
    def __init__(self, model: SheetGridTableModel, default_columns: list[int] | None = None, parent=None) -> None:
        super().__init__(parent)
        self._model = model
        self._result_columns: list[int] = []

        self.setWindowTitle("조합 열 추가")
        self.resize(760, 560)

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        title = QLabel("여러 열 값을 이어붙여 새 열 만들기")
        title.setStyleSheet("font-size:15px; font-weight:700;")
        root.addWidget(title)

        desc = QLabel(
            "예: A, B, D, E를 순서대로 선택하면 각 행에서 해당 칸 값을 한 칸 띄어쓰기하며 이어붙여 새 열에 저장합니다.\n"
            "헤더 1행은 새 열 제목으로 사용되고, 실제 조합은 2행부터 생성됩니다."
        )
        desc.setStyleSheet("color:#6b7280;")
        desc.setWordWrap(True)
        root.addWidget(desc)

        form = QGridLayout()
        form.setHorizontalSpacing(8)
        form.setVerticalSpacing(8)

        self.column_title_input = QLineEdit()
        self.column_title_input.setPlaceholderText("새 열 제목")
        self.column_title_input.setText("카카오검색명")

        self.separator_input = QLineEdit()
        self.separator_input.setPlaceholderText("구분자")
        self.separator_input.setText(" ")

        form.addWidget(QLabel("새 열 제목"), 0, 0)
        form.addWidget(self.column_title_input, 0, 1)
        form.addWidget(QLabel("값 구분자"), 1, 0)
        form.addWidget(self.separator_input, 1, 1)
        root.addLayout(form)

        list_row = QHBoxLayout()
        list_row.setSpacing(10)

        left_box = QVBoxLayout()
        left_box.addWidget(QLabel("사용 가능한 열"))
        self.available_list = QListWidget()
        left_box.addWidget(self.available_list, 1)
        list_row.addLayout(left_box, 1)

        middle_box = QVBoxLayout()
        middle_box.addStretch(1)
        self.btn_add_selected = QPushButton(">")
        self.btn_remove_selected = QPushButton("<")
        self.btn_add_all = QPushButton(">>")
        self.btn_clear_all = QPushButton("<<")
        for btn in [self.btn_add_selected, self.btn_remove_selected, self.btn_add_all, self.btn_clear_all]:
            btn.setFixedWidth(56)
            middle_box.addWidget(btn)
        middle_box.addStretch(1)
        list_row.addLayout(middle_box)

        right_box = QVBoxLayout()
        right_box.addWidget(QLabel("조합 순서"))
        self.selected_list = QListWidget()
        right_box.addWidget(self.selected_list, 1)

        order_btn_row = QHBoxLayout()
        self.btn_move_up = QPushButton("위로")
        self.btn_move_down = QPushButton("아래로")
        order_btn_row.addWidget(self.btn_move_up)
        order_btn_row.addWidget(self.btn_move_down)
        right_box.addLayout(order_btn_row)

        self.preview_label = QLabel("미리보기: ")
        self.preview_label.setStyleSheet("color:#6b7280;")
        self.preview_label.setWordWrap(True)
        right_box.addWidget(self.preview_label)

        list_row.addLayout(right_box, 1)
        root.addLayout(list_row, 1)

        bottom = QHBoxLayout()
        self.btn_cancel = QPushButton("취소")
        self.btn_apply = QPushButton("추가")
        bottom.addStretch(1)
        bottom.addWidget(self.btn_cancel)
        bottom.addWidget(self.btn_apply)
        root.addLayout(bottom)

        self.btn_add_selected.clicked.connect(self._add_selected_columns)
        self.btn_remove_selected.clicked.connect(self._remove_selected_columns)
        self.btn_add_all.clicked.connect(self._add_all_columns)
        self.btn_clear_all.clicked.connect(self._clear_all_columns)
        self.btn_move_up.clicked.connect(lambda: self._move_selected_item(-1))
        self.btn_move_down.clicked.connect(lambda: self._move_selected_item(1))
        self.btn_cancel.clicked.connect(self.reject)
        self.btn_apply.clicked.connect(self._accept_if_valid)
        self.available_list.itemDoubleClicked.connect(lambda _item: self._add_selected_columns())
        self.selected_list.itemDoubleClicked.connect(lambda _item: self._remove_selected_columns())
        self.selected_list.currentRowChanged.connect(lambda _row: self._update_preview())
        self.separator_input.textChanged.connect(lambda _text: self._update_preview())
        self.column_title_input.textChanged.connect(lambda _text: self._update_preview())

        self._populate_lists(default_columns or [])
        self._update_preview()

    def _column_item_text(self, col: int) -> str:
        header = self._model.index(0, col).data(Qt.DisplayRole) or ""
        sample = self._model.index(1, col).data(Qt.DisplayRole) or ""
        header_disp = str(header).strip() or "(제목없음)"
        sample_disp = str(sample).strip() or "(빈값)"
        return f"{get_column_letter(col + 1)} | {header_disp} | 예시: {sample_disp}"

    def _make_item(self, col: int) -> QListWidgetItem:
        item = QListWidgetItem(self._column_item_text(col))
        item.setData(Qt.UserRole, col)
        return item

    def _populate_lists(self, default_columns: list[int]) -> None:
        default_set = set(default_columns)
        col_count = self._model.columnCount()
        for col in range(col_count):
            item = self._make_item(col)
            if col in default_set:
                self.selected_list.addItem(item)
            else:
                self.available_list.addItem(item)
        self._sort_available()

    def _sort_available(self) -> None:
        items = []
        while self.available_list.count():
            item = self.available_list.takeItem(0)
            items.append(item)
        items.sort(key=lambda it: int(it.data(Qt.UserRole)))
        for item in items:
            self.available_list.addItem(item)

    def _take_selected_items(self, widget: QListWidget) -> list[QListWidgetItem]:
        rows = sorted({widget.row(item) for item in widget.selectedItems()}, reverse=True)
        taken = []
        for row in rows:
            item = widget.takeItem(row)
            if item is not None:
                taken.append(item)
        taken.reverse()
        return taken

    def _add_selected_columns(self) -> None:
        items = self._take_selected_items(self.available_list)
        for item in items:
            self.selected_list.addItem(item)
        self._update_preview()

    def _remove_selected_columns(self) -> None:
        items = self._take_selected_items(self.selected_list)
        for item in items:
            self.available_list.addItem(item)
        self._sort_available()
        self._update_preview()

    def _add_all_columns(self) -> None:
        while self.available_list.count():
            self.selected_list.addItem(self.available_list.takeItem(0))
        self._update_preview()

    def _clear_all_columns(self) -> None:
        while self.selected_list.count():
            self.available_list.addItem(self.selected_list.takeItem(0))
        self._sort_available()
        self._update_preview()

    def _move_selected_item(self, direction: int) -> None:
        row = self.selected_list.currentRow()
        if row < 0:
            return
        new_row = row + direction
        if new_row < 0 or new_row >= self.selected_list.count():
            return
        item = self.selected_list.takeItem(row)
        self.selected_list.insertItem(new_row, item)
        self.selected_list.setCurrentRow(new_row)
        self._update_preview()

    def _selected_columns(self) -> list[int]:
        cols: list[int] = []
        for i in range(self.selected_list.count()):
            item = self.selected_list.item(i)
            cols.append(int(item.data(Qt.UserRole)))
        return cols

    def _build_preview_text(self) -> str:
        cols = self._selected_columns()
        if not cols:
            return "미리보기: 선택된 열 없음"
        separator = self.separator_input.text()
        sample_values: list[str] = []
        for col in cols:
            value = self._model.index(1, col).data(Qt.DisplayRole) or ""
            value = str(value).strip()
            if value:
                sample_values.append(value)
        preview = separator.join(sample_values) if sample_values else ""
        title = (self.column_title_input.text() or "").strip() or "카카오검색명"
        return f"미리보기: [{title}] {preview}"

    def _update_preview(self) -> None:
        self.preview_label.setText(self._build_preview_text())

    def _accept_if_valid(self) -> None:
        cols = self._selected_columns()
        if not cols:
            QMessageBox.information(self, "안내", "조합에 사용할 열을 하나 이상 선택하세요.")
            return
        self._result_columns = cols
        self.accept()

    def get_result(self) -> tuple[str, str, list[int]]:
        title = (self.column_title_input.text() or "").strip() or "카카오검색명"
        separator = self.separator_input.text()
        return title, separator, list(self._result_columns)


class ExcelEditorDialog(QDialog):
    def __init__(self, grid: WorkbookGrid, parent=None) -> None:
        super().__init__(parent)
        self._grid = grid
        self._last_find_anchor: tuple[int, int] | None = None
        self._warned_value_only = False
        self._current_save_path = suggest_value_only_save_path(grid.source_path)
        self._undo_stack: list[EditCommand] = []
        self._redo_stack: list[EditCommand] = []
        self._column_focus_source = 0

        self.setWindowTitle(f"엑셀 미리보기/편집 - {Path(grid.source_path).name}")
        self.resize(1340, 860)

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        title = QLabel("엑셀 미리보기/편집")
        title.setStyleSheet("font-size:16px; font-weight:700;")
        root.addWidget(title)

        info = QLabel(
            f"원본 파일: {grid.source_path}  |  값 기준 경량 편집 모드 (서식/병합/매크로/수식 계산값 보존 안 함)"
        )
        info.setStyleSheet("color:#6b7280;")
        root.addWidget(info)

        self.chk_excel_format = QCheckBox("서식 보기")
        self.chk_excel_format.setChecked(False)
        root.addWidget(self.chk_excel_format, 0, Qt.AlignLeft)

        self.sample_header_value_labels: list[QLabel] = []
        self.sample_header_frame = self._build_sample_header_frame()
        self.sample_header_frame.hide()

        top_bar = QHBoxLayout()
        top_bar.setSpacing(8)
        top_bar.addWidget(QLabel("시트"))

        self.sheet_combo = QComboBox()
        self.sheet_combo.addItems(self._grid.sheetnames)
        top_bar.addWidget(self.sheet_combo)

        self.btn_undo = QPushButton("되돌리기")
        self.btn_redo = QPushButton("앞으로가기")
        self.btn_delete_rows = QPushButton("행 삭제")
        self.btn_delete_cols = QPushButton("열 삭제")
        self.btn_move_col_left = QPushButton("열 ← 이동")
        self.btn_move_col_right = QPushButton("열 → 이동")
        self.btn_add_combined_col = QPushButton("조합 열 추가")
        self.current_col_label = QLabel("현재 열: A")
        self.current_col_label.setStyleSheet("color:#6b7280;")
        self.btn_filter = QPushButton("열 필터")
        self.btn_clear_filters = QPushButton("필터 해제")
        self.filter_summary_label = QLabel("필터 없음")
        self.filter_summary_label.setStyleSheet("color:#6b7280;")
        self.btn_save = QPushButton("저장하기")
        self.btn_export = QPushButton("내보내기")
        self.btn_close = QPushButton("닫기")

        top_bar.addSpacing(10)
        top_bar.addWidget(self.btn_undo)
        top_bar.addWidget(self.btn_redo)
        top_bar.addWidget(self.btn_delete_rows)
        top_bar.addWidget(self.btn_delete_cols)
        top_bar.addWidget(self.btn_move_col_left)
        top_bar.addWidget(self.btn_move_col_right)
        top_bar.addWidget(self.btn_add_combined_col)
        top_bar.addWidget(self.current_col_label)
        top_bar.addWidget(self.btn_filter)
        top_bar.addWidget(self.btn_clear_filters)
        top_bar.addWidget(self.filter_summary_label)
        top_bar.addStretch(1)
        top_bar.addWidget(self.btn_save)
        top_bar.addWidget(self.btn_export)
        top_bar.addWidget(self.btn_close)
        root.addLayout(top_bar)

        find_bar = QHBoxLayout()
        find_bar.setSpacing(8)
        self.find_input = QLineEdit()
        self.find_input.setPlaceholderText("찾을 값")
        self.replace_input = QLineEdit()
        self.replace_input.setPlaceholderText("바꿀 값")
        self.chk_case = QCheckBox("대소문자 구분")
        self.btn_find = QPushButton("찾기")
        self.btn_replace_all = QPushButton("일괄 수정")

        find_bar.addWidget(QLabel("찾기"))
        find_bar.addWidget(self.find_input, 2)
        find_bar.addWidget(QLabel("바꾸기"))
        find_bar.addWidget(self.replace_input, 2)
        find_bar.addWidget(self.chk_case)
        find_bar.addWidget(self.btn_find)
        find_bar.addWidget(self.btn_replace_all)
        root.addLayout(find_bar)
        root.addWidget(self.sample_header_frame)

        self.table = QTableView()
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(False)
        self.table.setSelectionBehavior(QTableView.SelectItems)
        self.table.setSelectionMode(QTableView.ExtendedSelection)
        self.table.horizontalHeader().setDefaultSectionSize(120)
        self.table.verticalHeader().setDefaultSectionSize(26)
        self.table.horizontalHeader().setMinimumSectionSize(60)
        self.table.verticalHeader().setMinimumSectionSize(24)
        root.addWidget(self.table, 1)

        self.status_label = QLabel("준비 완료")
        self.status_label.setStyleSheet("color:#6b7280;")
        root.addWidget(self.status_label)

        self.model = SheetGridTableModel(self._grid.sheets[0], on_cell_edited=self._on_cell_edited)
        self.proxy = ColumnValueFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)
        self.table.setModel(self.proxy)

        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        self.chk_excel_format.toggled.connect(self._on_toggle_excel_format_view)
        self.btn_undo.clicked.connect(self._undo)
        self.btn_redo.clicked.connect(self._redo)
        self.btn_delete_rows.clicked.connect(self._delete_selected_rows)
        self.btn_delete_cols.clicked.connect(self._delete_selected_cols)
        self.btn_move_col_left.clicked.connect(lambda: self._move_current_column(-1))
        self.btn_move_col_right.clicked.connect(lambda: self._move_current_column(1))
        self.btn_add_combined_col.clicked.connect(self._add_combined_column)
        self.btn_filter.clicked.connect(self._open_filter_dialog)
        self.btn_clear_filters.clicked.connect(self._clear_filters)
        self.btn_find.clicked.connect(self._find_next)
        self.btn_replace_all.clicked.connect(self._replace_all)
        self.btn_save.clicked.connect(self._save_to_default)
        self.btn_export.clicked.connect(self._export_as)
        self.btn_close.clicked.connect(self.accept)
        self.find_input.returnPressed.connect(self._find_next)
        self.replace_input.returnPressed.connect(self._replace_all)

        QShortcut(QKeySequence("Ctrl+F"), self, activated=self._focus_find)
        QShortcut(QKeySequence("Ctrl+H"), self, activated=self._focus_replace)
        QShortcut(QKeySequence("F3"), self, activated=self._find_next)
        QShortcut(QKeySequence("Ctrl+Z"), self, activated=self._undo)
        QShortcut(QKeySequence("Ctrl+Y"), self, activated=self._redo)
        QShortcut(QKeySequence("Ctrl+Shift+Z"), self, activated=self._redo)
        QShortcut(QKeySequence("Ctrl+Shift+Left"), self, activated=lambda: self._move_current_column(-1))
        QShortcut(QKeySequence("Ctrl+Shift+Right"), self, activated=lambda: self._move_current_column(1))

        self.table.horizontalHeader().sectionClicked.connect(self._on_header_section_clicked)
        self.table.horizontalHeader().sectionResized.connect(self._sync_sample_header_widths)
        self.table.horizontalScrollBar().valueChanged.connect(self._sync_sample_header_scroll)
        self.table.verticalHeader().sectionResized.connect(self._sync_sample_header_widths)
        if self.table.selectionModel():
            self.table.selectionModel().currentChanged.connect(self._on_current_changed)

        self._update_history_buttons()
        self._update_filter_summary()
        self._sync_sample_header_widths()
        self._set_status(self._sheet_summary_text(self._grid.sheets[0]))

    def _clone_rows(self, rows: list[list[str]]) -> list[list[str]]:
        return [list(row) for row in rows]

    def _current_sheet_index(self) -> int:
        idx = self.sheet_combo.currentIndex()
        return idx if idx >= 0 else 0

    def _current_source_column(self) -> int:
        current = self.table.currentIndex()
        if current.isValid():
            source = self.proxy.mapToSource(current)
            if source.isValid():
                return source.column()
        return self._column_focus_source

    def _push_command(self, command: EditCommand) -> None:
        self._undo_stack.append(command)
        self._redo_stack.clear()
        self._update_history_buttons()

    def _update_history_buttons(self) -> None:
        self.btn_undo.setEnabled(len(self._undo_stack) > 0)
        self.btn_redo.setEnabled(len(self._redo_stack) > 0)

    def _update_filter_summary(self) -> None:
        active = self.proxy.active_filters()
        self.btn_clear_filters.setEnabled(len(active) > 0)
        if not active:
            self.filter_summary_label.setText("필터 없음")
            return
        labels = [get_column_letter(col + 1) for col in sorted(active.keys())]
        summary = ", ".join(labels[:4])
        if len(labels) > 4:
            summary += " ..."
        self.filter_summary_label.setText(f"필터 적용: {summary}")

    def _refresh_after_data_change(self) -> None:
        self.proxy.invalidateFilter()
        self.proxy.prune_invalid_columns(self.model.columnCount())
        max_col = max(0, self.model.columnCount() - 1)
        self._column_focus_source = min(self._column_focus_source, max_col)
        self._update_filter_summary()
        self._update_current_col_label()
        self._sync_sample_header_widths()

    def _on_cell_edited(self, row: int, col: int, old_value: str, new_value: str) -> None:
        self._push_command(
            CellEditCommand(
                sheet_index=self._current_sheet_index(),
                row=row,
                col=col,
                old_value=old_value,
                new_value=new_value,
            )
        )
        self._refresh_after_data_change()
        self._set_status(f"셀 수정: {get_column_letter(col + 1)}{row + 1}")

    def _focus_find(self) -> None:
        self.find_input.setFocus()
        self.find_input.selectAll()

    def _focus_replace(self) -> None:
        self.replace_input.setFocus()
        self.replace_input.selectAll()

    def _build_sample_header_frame(self) -> QFrame:
        frame = QFrame()
        frame.setStyleSheet("QFrame { background:transparent; border:none; }")
        outer = QHBoxLayout(frame)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self.sample_row_header_label = QLabel("서식")
        self.sample_row_header_label.setAlignment(Qt.AlignCenter)
        self.sample_row_header_label.setStyleSheet(
            "background:#f3f4f6; border:1px solid #d0d7e2; border-right:none; min-height:28px; font-weight:600; color:#374151;"
        )
        self.sample_row_header_label.setFixedWidth(44)
        outer.addWidget(self.sample_row_header_label)

        self.sample_header_scroll = QScrollArea()
        self.sample_header_scroll.setWidgetResizable(False)
        self.sample_header_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sample_header_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sample_header_scroll.setFrameShape(QFrame.NoFrame)
        self.sample_header_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.sample_header_scroll.setFixedHeight(30)

        self.sample_header_content = QFrame()
        self.sample_header_content.setStyleSheet("QFrame { background:transparent; border:none; }")
        row = QHBoxLayout(self.sample_header_content)
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(0)

        self.sample_header_value_labels.clear()
        for idx, header_text in enumerate(EXCEL_SAMPLE_HEADERS):
            value = QLabel(header_text)
            value.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            value.setStyleSheet(
                "background:#ffffff; border:1px solid #d0d7e2; border-left:none; min-height:28px; padding:0 8px; font-weight:600; color:#111827;"
            )
            value.setTextInteractionFlags(Qt.TextSelectableByMouse)
            value.setFixedWidth(120)
            self.sample_header_value_labels.append(value)
            row.addWidget(value)

        self.sample_header_scroll.setWidget(self.sample_header_content)
        outer.addWidget(self.sample_header_scroll, 1)
        return frame

    def _on_toggle_excel_format_view(self, checked: bool) -> None:
        self.sample_header_frame.setVisible(checked)
        if checked:
            self._sync_sample_header_widths()

    def _sync_sample_header_widths(self, *args) -> None:
        if not hasattr(self, "table"):
            return
        header = self.table.horizontalHeader()
        vertical_header = self.table.verticalHeader()
        if hasattr(self, "sample_row_header_label"):
            self.sample_row_header_label.setFixedWidth(max(44, vertical_header.width()))
        total_width = 0
        for idx, label in enumerate(self.sample_header_value_labels):
            if idx < header.count():
                width = header.sectionSize(idx)
            else:
                width = header.defaultSectionSize()
            width = max(80, width)
            label.setFixedWidth(width)
            total_width += width
        if hasattr(self, "sample_header_content"):
            self.sample_header_content.setFixedWidth(total_width)
        if hasattr(self, "sample_header_scroll") and hasattr(self.table, "horizontalScrollBar"):
            self.sample_header_scroll.horizontalScrollBar().setValue(self.table.horizontalScrollBar().value())

    def _sync_sample_header_scroll(self, value: int) -> None:
        if hasattr(self, "sample_header_scroll"):
            self.sample_header_scroll.horizontalScrollBar().setValue(value)

    def _on_current_changed(self, current: QModelIndex, previous: QModelIndex) -> None:
        if current.isValid():
            source = self.proxy.mapToSource(current)
            if source.isValid():
                self._column_focus_source = source.column()
        self._update_current_col_label()

    def _on_header_section_clicked(self, logical_index: int) -> None:
        self._column_focus_source = logical_index
        if self.proxy.rowCount() > 0:
            idx = self.proxy.index(0, logical_index)
            if idx.isValid():
                self.table.setCurrentIndex(idx)
                selection_model = self.table.selectionModel()
                if selection_model:
                    selection_model.clearSelection()
                    selection_model.select(idx, QItemSelectionModel.ClearAndSelect)
        self._update_current_col_label()

    def _update_current_col_label(self) -> None:
        col = max(0, self._current_source_column())
        self.current_col_label.setText(f"현재 열: {get_column_letter(col + 1)}")

    def _move_current_column(self, direction: int) -> None:
        source_col = self._current_source_column()
        target_col = source_col + direction
        col_count = self.model.columnCount()
        if source_col < 0 or source_col >= col_count:
            QMessageBox.information(self, "안내", "이동할 열을 먼저 선택하세요.")
            return
        if target_col < 0 or target_col >= col_count:
            return

        row = 0
        current = self.table.currentIndex()
        if current.isValid():
            source_idx = self.proxy.mapToSource(current)
            if source_idx.isValid():
                row = source_idx.row()

        before_rows = self.model.get_rows_copy()
        moved = self.model.move_column(source_col, target_col)
        if not moved:
            return
        after_rows = self.model.get_rows_copy()
        self._record_snapshot_command(before_rows, after_rows, "열 이동")
        self._clear_filters(silent=True)
        self._refresh_after_data_change()
        self._column_focus_source = target_col
        self._last_find_anchor = None
        self._select_cell(max(0, row), target_col)
        self._update_current_col_label()
        self._set_status(
            f"열 이동 완료: {get_column_letter(source_col + 1)} → {get_column_letter(target_col + 1)} (필터 해제)"
        )

    def _add_combined_column(self) -> None:
        default_cols = self._selected_columns_for_combine()
        dialog = CombinedColumnDialog(self.model, default_columns=default_cols, parent=self)
        if dialog.exec() != QDialog.Accepted:
            return

        header_title, separator, source_cols = dialog.get_result()
        if not source_cols:
            return

        before_rows = self.model.get_rows_copy()
        new_col = self.model.add_combined_column(source_cols, header_title, separator)
        after_rows = self.model.get_rows_copy()
        self._record_snapshot_command(before_rows, after_rows, "조합 열 추가")
        self._clear_filters(silent=True)
        self._refresh_after_data_change()
        self._column_focus_source = new_col
        self._last_find_anchor = None
        target_row = 1 if self.model.rowCount() > 1 else 0
        self._select_cell(target_row, new_col)

        labels = ", ".join(get_column_letter(c + 1) for c in source_cols)
        self._set_status(f"조합 열 추가 완료: {get_column_letter(new_col + 1)}열 / 원본 {labels}")
        QMessageBox.information(
            self,
            "완료",
            f"조합 열이 추가되었습니다.\n\n새 열: {get_column_letter(new_col + 1)} ({header_title})\n원본 열: {labels}"
        )

    def _open_filter_dialog(self) -> None:
        dialog = ColumnFilterDialog(
            model=self.model,
            active_filters=self.proxy.active_filters(),
            initial_column=self._current_source_column(),
            parent=self,
        )
        if dialog.exec() != QDialog.Accepted:
            return

        new_filters = dialog.get_filters()
        self.proxy.clear_all_filters()
        for col, values in new_filters.items():
            self.proxy.set_allowed_values(col, values)
        self._last_find_anchor = None
        self._update_filter_summary()
        self._set_status("열 필터 적용 완료")

    def _clear_filters(self, silent: bool = False) -> None:
        if not self.proxy.has_active_filters():
            return
        self.proxy.clear_all_filters()
        self._last_find_anchor = None
        self._update_filter_summary()
        if not silent:
            self._set_status("필터 해제 완료")

    def _on_sheet_changed(self, index: int) -> None:
        if index < 0 or index >= len(self._grid.sheets):
            return
        self.model.set_sheet(self._grid.sheets[index])
        self._column_focus_source = 0
        self._clear_filters(silent=True)
        self._last_find_anchor = None
        self._update_current_col_label()
        self._set_status(self._sheet_summary_text(self._grid.sheets[index]))

    def _sheet_summary_text(self, sheet: SheetGrid) -> str:
        return f"시트: {sheet.name} / {self.model.rowCount()}행 × {self.model.columnCount()}열"

    def _set_status(self, text: str) -> None:
        self.status_label.setText(text)

    def _selected_columns_for_combine(self) -> list[int]:
        selection_model = self.table.selectionModel()
        if not selection_model:
            return []
        indexes = selection_model.selectedIndexes()
        cols = sorted({self.proxy.mapToSource(idx).column() for idx in indexes if self.proxy.mapToSource(idx).isValid()})
        if cols:
            return cols
        current = self.table.currentIndex()
        if not current.isValid():
            return []
        source = self.proxy.mapToSource(current)
        return [source.column()] if source.isValid() else []

    def _selected_rows(self) -> list[int]:
        selection_model = self.table.selectionModel()
        if not selection_model:
            return []
        indexes = selection_model.selectedIndexes()
        rows = sorted({self.proxy.mapToSource(idx).row() for idx in indexes if self.proxy.mapToSource(idx).isValid()})
        if rows:
            return rows
        current = self.table.currentIndex()
        if not current.isValid():
            return []
        source = self.proxy.mapToSource(current)
        return [source.row()] if source.isValid() else []

    def _selected_cols(self) -> list[int]:
        selection_model = self.table.selectionModel()
        if not selection_model:
            return []
        indexes = selection_model.selectedIndexes()
        cols = sorted({self.proxy.mapToSource(idx).column() for idx in indexes if self.proxy.mapToSource(idx).isValid()})
        if cols:
            return cols
        current = self.table.currentIndex()
        if not current.isValid():
            return []
        source = self.proxy.mapToSource(current)
        return [source.column()] if source.isValid() else []

    def _ensure_sheet_active(self, sheet_index: int) -> None:
        if self.sheet_combo.currentIndex() != sheet_index:
            self.sheet_combo.setCurrentIndex(sheet_index)

    def _select_cell(self, row: int, col: int) -> bool:
        source_index = self.model.index(max(0, row), max(0, col))
        if not source_index.isValid():
            return False
        index = self.proxy.mapFromSource(source_index)
        if not index.isValid():
            return False
        self.table.setCurrentIndex(index)
        self.table.scrollTo(index, QTableView.PositionAtCenter)
        selection_model = self.table.selectionModel()
        if selection_model:
            selection_model.clearSelection()
            selection_model.select(index, QItemSelectionModel.ClearAndSelect)
        self._column_focus_source = col
        self._update_current_col_label()
        self._last_find_anchor = (row, col)
        return True

    def _record_snapshot_command(self, before_rows: list[list[str]], after_rows: list[list[str]], label: str) -> None:
        if before_rows == after_rows:
            return
        self._push_command(
            SheetSnapshotCommand(
                sheet_index=self._current_sheet_index(),
                before_rows=self._clone_rows(before_rows),
                after_rows=self._clone_rows(after_rows),
                label=label,
            )
        )

    def _apply_command(self, command: EditCommand, use_after_state: bool) -> None:
        self._ensure_sheet_active(command.sheet_index)
        self.model.set_history_suspended(True)
        try:
            if isinstance(command, CellEditCommand):
                value = command.new_value if use_after_state else command.old_value
                index = self.model.index(command.row, command.col)
                self.model.setData(index, value, Qt.EditRole)
                self._refresh_after_data_change()
                self._select_cell(command.row, command.col)
            else:
                rows = command.after_rows if use_after_state else command.before_rows
                self.model.replace_rows(self._clone_rows(rows))
                self._refresh_after_data_change()
                self._clear_filters(silent=True)
                self._last_find_anchor = None
                if self.model.rowCount() > 0 and self.model.columnCount() > 0:
                    self._select_cell(0, 0)
        finally:
            self.model.set_history_suspended(False)

    def _undo(self) -> None:
        if not self._undo_stack:
            return
        command = self._undo_stack.pop()
        self._apply_command(command, use_after_state=False)
        self._redo_stack.append(command)
        self._update_history_buttons()
        label = "셀 수정" if isinstance(command, CellEditCommand) else command.label
        self._set_status(f"되돌리기 완료: {label}")

    def _redo(self) -> None:
        if not self._redo_stack:
            return
        command = self._redo_stack.pop()
        self._apply_command(command, use_after_state=True)
        self._undo_stack.append(command)
        self._update_history_buttons()
        label = "셀 수정" if isinstance(command, CellEditCommand) else command.label
        self._set_status(f"앞으로가기 완료: {label}")

    def _delete_selected_rows(self) -> None:
        rows = self._selected_rows()
        if not rows:
            QMessageBox.information(self, "안내", "삭제할 행을 선택하세요.")
            return
        ok = QMessageBox.question(
            self,
            "행 삭제",
            f"선택한 {len(rows)}개 행을 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if ok != QMessageBox.Yes:
            return
        before_rows = self.model.get_rows_copy()
        deleted = self.model.delete_rows(rows)
        after_rows = self.model.get_rows_copy()
        self._record_snapshot_command(before_rows, after_rows, "행 삭제")
        self._refresh_after_data_change()
        self._last_find_anchor = None
        self._set_status(f"행 삭제 완료: {deleted}개")

    def _delete_selected_cols(self) -> None:
        cols = self._selected_cols()
        if not cols:
            QMessageBox.information(self, "안내", "삭제할 열을 선택하세요.")
            return
        labels = ", ".join(get_column_letter(c + 1) for c in cols[:8])
        if len(cols) > 8:
            labels += " ..."
        ok = QMessageBox.question(
            self,
            "열 삭제",
            f"선택한 {len(cols)}개 열({labels})을 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if ok != QMessageBox.Yes:
            return
        before_rows = self.model.get_rows_copy()
        deleted = self.model.delete_columns(cols)
        after_rows = self.model.get_rows_copy()
        self._record_snapshot_command(before_rows, after_rows, "열 삭제")
        self.proxy.shift_after_deleted_columns(cols)
        self._update_filter_summary()
        self._last_find_anchor = None
        self._set_status(f"열 삭제 완료: {deleted}개")

    def _find_next(self) -> None:
        needle = self.find_input.text().strip()
        if not needle:
            QMessageBox.information(self, "안내", "찾을 값을 입력하세요.")
            return

        found = self.model.find_next(
            needle=needle,
            start_after=self._last_find_anchor,
            case_sensitive=self.chk_case.isChecked(),
        )
        if found is None:
            QMessageBox.information(self, "찾기", "일치하는 값을 찾지 못했습니다.")
            self._last_find_anchor = None
            return

        row, col = found
        if self._select_cell(row, col):
            self._set_status(f"찾음: {get_column_letter(col + 1)}{row + 1}")
            return

        self._last_find_anchor = (row, col)
        if self.proxy.has_active_filters():
            ok = QMessageBox.question(
                self,
                "필터로 숨김",
                "찾은 셀이 현재 필터 조건으로 숨겨져 있습니다.\n필터를 모두 해제하고 이동하시겠습니까?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if ok == QMessageBox.Yes:
                self._clear_filters(silent=True)
                if self._select_cell(row, col):
                    self._set_status(f"찾음: {get_column_letter(col + 1)}{row + 1}")
                    return
        QMessageBox.information(self, "찾기", "찾은 셀이 현재 화면에 표시되지 않습니다.")

    def _replace_all(self) -> None:
        needle = self.find_input.text()
        if needle == "":
            QMessageBox.information(self, "안내", "찾을 값을 입력하세요.")
            return
        replacement = self.replace_input.text()
        before_rows = self.model.get_rows_copy()
        changed = self.model.replace_all(needle, replacement, self.chk_case.isChecked())
        after_rows = self.model.get_rows_copy()
        self._record_snapshot_command(before_rows, after_rows, "일괄 수정")
        self._refresh_after_data_change()
        self._last_find_anchor = None
        self._set_status(f"일괄 수정 완료: {changed}개 셀 변경")
        QMessageBox.information(self, "완료", f"일괄 수정 완료: {changed}개 셀 변경")

    def _confirm_value_only_save(self) -> bool:
        if self._warned_value_only:
            return True
        msg = (
            "현재 편집기는 값 기준 경량 편집 모드입니다.\n\n"
            "저장/내보내기 시 서식, 병합, 매크로, 수식 계산 결과는 보존되지 않고\n"
            "편집된 셀 값만 새 Excel 구조로 저장됩니다.\n\n"
            "계속 진행하시겠습니까?"
        )
        ok = QMessageBox.question(self, "저장 방식 안내", msg, QMessageBox.Yes | QMessageBox.No)
        if ok == QMessageBox.Yes:
            self._warned_value_only = True
            return True
        return False

    def _save_to_default(self) -> None:
        if not self._confirm_value_only_save():
            return
        target = self._current_save_path
        try:
            save_workbook_grid_to_xlsx(self._grid, target)
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 실패:\n{e}")
            return
        self._set_status(f"저장 완료: {target}")
        QMessageBox.information(self, "완료", f"저장 완료\n\n{target}")

    def _export_as(self) -> None:
        if not self._confirm_value_only_save():
            return
        default_name = f"{Path(self._current_save_path).stem}.xlsx"
        try:
            path = pick_save_file(
                title="편집 데이터 내보내기",
                filters=[Filter("Excel Files", "*.xlsx"), Filter("All Files", "*.*")],
                default_filename=default_name,
                default_ext="xlsx",
            )
        except Exception as e:
            QMessageBox.critical(self, "오류", f"저장 위치 선택기 실행 실패:\n{e}")
            return
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        try:
            save_workbook_grid_to_xlsx(self._grid, path)
        except Exception as e:
            QMessageBox.critical(self, "오류", f"내보내기 실패:\n{e}")
            return
        self._current_save_path = path
        self._set_status(f"내보내기 완료: {path}")
        QMessageBox.information(self, "완료", f"내보내기 완료\n\n{path}")
