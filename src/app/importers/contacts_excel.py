from __future__ import annotations

from dataclasses import dataclass
from typing import List, Tuple
from openpyxl import load_workbook


REQUIRED_HEADERS = ["사번", "이름", "전화번호", "대리점명", "지사명"]


@dataclass
class ImportResult:
    rows: List[Tuple[str, str, str, str, str]]  # (emp_id, name, phone, agency, branch)
    skipped: int
    errors: List[str]


def import_contacts_xlsx(path: str) -> ImportResult:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active

    # 1행 헤더 읽기
    header_cells = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    headers = [str(x).strip() if x is not None else "" for x in header_cells]

    # 필요한 헤더 위치 매핑
    idx = {}
    for h in REQUIRED_HEADERS:
        if h not in headers:
            return ImportResult(rows=[], skipped=0, errors=[f"필수 컬럼 누락: {h}"])
        idx[h] = headers.index(h) + 1  # 1-based column

    rows: List[Tuple[str, str, str, str, str]] = []
    errors: List[str] = []
    skipped = 0

    # 2행부터 데이터
    for r in range(2, ws.max_row + 1):
        emp_id = ws.cell(row=r, column=idx["사번"]).value
        name = ws.cell(row=r, column=idx["이름"]).value
        phone = ws.cell(row=r, column=idx["전화번호"]).value
        agency = ws.cell(row=r, column=idx["대리점명"]).value
        branch = ws.cell(row=r, column=idx["지사명"]).value

        emp_id = "" if emp_id is None else str(emp_id).strip()
        name = "" if name is None else str(name).strip()
        phone = "" if phone is None else str(phone).strip()
        agency = "" if agency is None else str(agency).strip()
        branch = "" if branch is None else str(branch).strip()

        # 최소 검증(사번/이름 필수)
        if not emp_id or not name:
            skipped += 1
            continue

        rows.append((emp_id, name, phone, agency, branch))

    return ImportResult(rows=rows, skipped=skipped, errors=errors)
