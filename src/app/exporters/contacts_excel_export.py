from __future__ import annotations

import os
import tempfile
from datetime import datetime
from typing import Iterable, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = ["사번", "이름", "전화번호", "대리점명", "지사명"]


def _apply_sheet_style(ws) -> None:
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")  # 진한 파랑
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(HEADERS)
    ws.freeze_panes = "A2"

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20

    # Column widths
    widths = [12, 12, 18, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _atomic_save_workbook(wb: Workbook, path: str) -> None:
    """
    ✅ Windows에서 '엑셀 파일이 열려있음/잠금' 등으로 저장 실패하는 문제를 줄이기 위한 원자적 저장
    - 같은 폴더에 임시파일로 저장 -> 성공 시 os.replace로 교체
    """
    path = os.path.abspath(path)
    folder = os.path.dirname(path) or "."
    os.makedirs(folder, exist_ok=True)

    fd, tmp_path = tempfile.mkstemp(prefix=".__tmp_", suffix=".xlsx", dir=folder)
    os.close(fd)

    try:
        wb.save(tmp_path)
        # target이 존재하든 말든 교체(동일 드라이브에서 원자적)
        os.replace(tmp_path, path)
    except Exception:
        # 실패 시 임시파일 정리
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        raise


def create_template_xlsx(path: str) -> None:
    """
    샘플 서식(헤더 + 안내 시트) 생성
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "대상자"
    _apply_sheet_style(ws)

    # 예시 2행 (사용자 이해용)
    ws.append(["1001", "홍길동", "01011112222", "강남대리점", "서울"])
    ws.append(["1002", "김영희", "01033334444", "성수대리점", "서울"])

    # 안내 시트
    ws2 = wb.create_sheet("안내")
    ws2["A1"] = "입력 규칙"
    ws2["A1"].font = Font(bold=True)
    ws2["A3"] = "1) 1행 헤더는 수정하지 마세요. (사번/이름/전화번호/대리점명/지사명)"
    ws2["A4"] = "2) 사번, 이름은 필수입니다. 누락 시 Import에서 스킵됩니다."
    ws2["A5"] = "3) 파일 형식은 .xlsx만 지원합니다."
    ws2["A6"] = f"4) 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    _atomic_save_workbook(wb, path)


def export_contacts_xlsx(path: str, rows: Iterable[Tuple[str, str, str, str, str]]) -> None:
    """
    rows: (emp_id, name, phone, agency, branch)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "대상자"
    _apply_sheet_style(ws)

    for r in rows:
        ws.append(list(r))

    _atomic_save_workbook(wb, path)
