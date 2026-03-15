# FILE: src/backend/integrations/excel/workbook_editor_io.py
from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


SUPPORTED_EDITOR_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


@dataclass
class SheetGrid:
    name: str
    rows: list[list[str]]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def col_count(self) -> int:
        if not self.rows:
            return 0
        return max((len(r) for r in self.rows), default=0)

    def ensure_rectangular(self) -> None:
        width = max(1, self.col_count)
        if not self.rows:
            self.rows = [[""]]
            return
        for row in self.rows:
            if len(row) < width:
                row.extend([""] * (width - len(row)))


@dataclass
class WorkbookGrid:
    source_path: str
    sheets: list[SheetGrid]

    @property
    def sheetnames(self) -> list[str]:
        return [s.name for s in self.sheets]


def is_supported_excel_editor_file(path: str) -> bool:
    return Path(path).suffix.lower() in SUPPORTED_EDITOR_EXTS


def load_workbook_grid(path: str) -> WorkbookGrid:
    ext = Path(path).suffix.lower()
    if ext not in SUPPORTED_EDITOR_EXTS:
        raise ValueError("지원 확장자: .xlsx, .xlsm, .xltx, .xltm")

    keep_vba = ext == ".xlsm"
    wb = load_workbook(filename=path, read_only=True, data_only=False, keep_vba=keep_vba)
    try:
        sheets: list[SheetGrid] = []
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            max_cols = 0
            for row in ws.iter_rows(values_only=True):
                values = [_cell_to_text(v) for v in row]
                while values and values[-1] == "":
                    values.pop()
                rows.append(values)
                if len(values) > max_cols:
                    max_cols = len(values)

            if not rows:
                rows = [[""]]
                max_cols = 1

            if max_cols <= 0:
                max_cols = 1

            normalized: list[list[str]] = []
            for row in rows:
                if len(row) < max_cols:
                    normalized.append(row + [""] * (max_cols - len(row)))
                else:
                    normalized.append(row)

            sheet = SheetGrid(name=ws.title, rows=normalized)
            sheet.ensure_rectangular()
            sheets.append(sheet)

        if not sheets:
            sheets.append(SheetGrid(name="Sheet1", rows=[[""]]))

        return WorkbookGrid(source_path=path, sheets=sheets)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def save_workbook_grid_to_xlsx(grid: WorkbookGrid, path: str) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet in grid.sheets:
        ws = wb.create_sheet(title=_safe_sheet_title(sheet.name, wb.sheetnames))
        sheet.ensure_rectangular()
        for row in sheet.rows:
            ws.append([_write_cell_value(v) for v in row])

    if not wb.worksheets:
        wb.create_sheet(title="Sheet1")

    _save_workbook_atomic(wb, path)


def suggest_value_only_save_path(source_path: str) -> str:
    src = Path(source_path)
    if src.suffix.lower() == ".xlsx":
        return str(src)
    return str(src.with_name(f"{src.stem}_edited.xlsx"))


def _save_workbook_atomic(wb: Workbook, path: str) -> None:
    path = os.path.abspath(path)
    folder = os.path.dirname(path) or "."
    os.makedirs(folder, exist_ok=True)

    fd, tmp_path = tempfile.mkstemp(prefix=".__tmp_", suffix=".xlsx", dir=folder)
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        raise


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def _write_cell_value(value: str):
    text = "" if value is None else str(value)
    return text if text != "" else None


def _safe_sheet_title(title: str, existing: list[str]) -> str:
    cleaned = (title or "Sheet").strip()
    cleaned = cleaned.replace("/", "_").replace("\\", "_").replace("*", "_").replace("?", "_")
    cleaned = cleaned.replace(":", "_").replace("[", "(").replace("]", ")")
    cleaned = cleaned[:31] or "Sheet"

    if cleaned not in existing:
        return cleaned

    base = cleaned[:28] or "Sheet"
    i = 1
    while True:
        candidate = f"{base}_{i}"[:31]
        if candidate not in existing:
            return candidate
        i += 1
