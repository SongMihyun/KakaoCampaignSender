from __future__ import annotations

import csv
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook


PreviewRow = dict[str, str]

CONTACT_IMPORT_KEYS = [
    "emp_id",
    "name",
    "customer_name",
    "customer_honorific",
    "customer_position",
    "agency",
    "branch",
    "phone",
    "customer_status",
    "tags",
    "memo2",
]

_HEADER_ALIASES: dict[str, set[str]] = {
    "emp_id": {"사번", "사원번호", "고객번호", "empid", "emp_id", "employeeid", "employee_id", "id"},
    "name": {"카카오톡검색명", "카톡검색명", "검색명", "이름", "성명", "name"},
    "customer_name": {"고객명", "고객이름", "수신자명", "실제이름", "customername", "customer_name"},
    "customer_honorific": {"호칭", "고객호칭", "honorific"},
    "customer_position": {"직책", "직함", "직위", "position", "title"},
    "agency": {"소속대리점", "소속", "대리점명", "대리점", "agency", "company"},
    "branch": {"지사명", "지사", "branch"},
    "phone": {"연락처", "전화번호", "전화", "휴대폰", "핸드폰", "휴대전화", "phone", "mobile", "tel"},
    "customer_status": {"상태", "고객상태", "status"},
    "tags": {"태그", "tag", "tags"},
    "memo2": {"메모", "비고", "memo", "note", "notes"},
}

_DOCX_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
_TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "euc-kr")
_SUPPORTED_EXTS = {".xlsx", ".xlsm", ".docx", ".txt", ".csv", ".tsv"}


@dataclass
class ImportResult:
    rows: list[PreviewRow]
    skipped: int
    errors: list[str]


def supported_contact_import_exts() -> set[str]:
    return set(_SUPPORTED_EXTS)


def is_supported_contact_import_file(path: str) -> bool:
    return Path(path).suffix.lower() in _SUPPORTED_EXTS


def import_contacts_file(path: str) -> ImportResult:
    ext = Path(path).suffix.lower()

    try:
        if ext in {".xlsx", ".xlsm"}:
            raw_rows = _read_xlsx_rows(path)
        elif ext == ".docx":
            raw_rows = _read_docx_rows(path)
        elif ext in {".txt", ".csv", ".tsv"}:
            raw_rows = _read_text_rows(path)
        else:
            return ImportResult(
                rows=[],
                skipped=0,
                errors=["지원하지 않는 파일 형식입니다. 지원 확장자: .xlsx, .xlsm, .docx, .txt, .csv, .tsv"],
            )
    except Exception as e:
        return ImportResult(rows=[], skipped=0, errors=[f"파일 파싱 실패: {e}"])

    return _build_import_result(raw_rows)


def import_contacts_text(text: str) -> ImportResult:
    try:
        raw_rows = _rows_from_text_blob(text)
    except Exception as e:
        return ImportResult(rows=[], skipped=0, errors=[f"붙여넣기 파싱 실패: {e}"])
    return _build_import_result(raw_rows)


def import_contacts_xlsx(path: str) -> ImportResult:
    return import_contacts_file(path)


def _read_xlsx_rows(path: str) -> list[list[str]]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    return [[_normalize_cell(v) for v in row] for row in ws.iter_rows(values_only=True)]


def _read_text_rows(path: str) -> list[list[str]]:
    text = _read_text_file(path)
    return _rows_from_text_blob(text)


def _rows_from_text_blob(text: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for raw_line in (text or "").splitlines():
        line = raw_line.rstrip("\r")
        if not line.strip():
            rows.append([])
            continue
        rows.append(_split_text_line(line))
    return rows


def _read_text_file(path: str) -> str:
    last_error: Exception | None = None
    for enc in _TEXT_ENCODINGS:
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError as e:
            last_error = e
            continue

    if last_error is not None:
        raise last_error

    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def _read_docx_rows(path: str) -> list[list[str]]:
    with zipfile.ZipFile(path) as zf:
        try:
            xml_bytes = zf.read("word/document.xml")
        except KeyError as e:
            raise ValueError("Word 문서에서 본문(word/document.xml)을 찾을 수 없습니다.") from e

    root = ET.fromstring(xml_bytes)
    body = root.find("w:body", _DOCX_NS)
    if body is None:
        return []

    rows: list[list[str]] = []
    for child in list(body):
        tag = _local_name(child.tag)
        if tag == "tbl":
            rows.extend(_extract_docx_table_rows(child))
        elif tag == "p":
            text = _extract_docx_paragraph_text(child)
            if text:
                rows.append(_split_text_line(text))
    return rows


def _extract_docx_table_rows(tbl_el: ET.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for tr in tbl_el.findall("w:tr", _DOCX_NS):
        cells: list[str] = []
        for tc in tr.findall("w:tc", _DOCX_NS):
            parts = [t.text or "" for t in tc.findall(".//w:t", _DOCX_NS)]
            cells.append(_cleanup_text("".join(parts)))
        rows.append(cells)
    return rows


def _extract_docx_paragraph_text(p_el: ET.Element) -> str:
    parts = [t.text or "" for t in p_el.findall(".//w:t", _DOCX_NS)]
    return _cleanup_text("".join(parts))


def _build_import_result(raw_rows: Sequence[Sequence[str]]) -> ImportResult:
    cleaned_rows = [_trim_row(row) for row in raw_rows]
    cleaned_rows = [row for row in cleaned_rows if row]

    if not cleaned_rows:
        return ImportResult(rows=[], skipped=0, errors=[])

    rows: list[PreviewRow] = []
    skipped = 0
    active_header_map: dict[str, int] | None = None

    for raw in cleaned_rows:
        detected_header = _detect_header_map(raw)
        if detected_header:
            active_header_map = detected_header
            continue

        rec = _row_to_record(raw, active_header_map)
        rec["name"] = _cleanup_name(rec["name"] or rec["customer_name"])
        rec["customer_name"] = _cleanup_name(rec["customer_name"] or rec["name"])
        rec["customer_honorific"] = rec["customer_honorific"] or "고객님"

        if not rec["name"]:
            skipped += 1
            continue

        rows.append(rec)

    return ImportResult(rows=rows, skipped=skipped, errors=[])


def _detect_header_map(row: Sequence[str]) -> dict[str, int] | None:
    mapped: dict[str, int] = {}

    for idx, cell in enumerate(row):
        key = _map_header(cell)
        if key and key not in mapped:
            mapped[key] = idx

    if not mapped:
        return None

    if "name" in mapped or "customer_name" in mapped:
        return mapped
    if len(mapped) >= 3 and 0 in mapped.values():
        return mapped
    return None


def _empty_record() -> PreviewRow:
    return {key: "" for key in CONTACT_IMPORT_KEYS}


def _row_to_record(row: Sequence[str], header_map: dict[str, int] | None) -> PreviewRow:
    if header_map:
        rec = _empty_record()
        for key in CONTACT_IMPORT_KEYS:
            rec[key] = _get_cell(row, header_map.get(key))
        return rec

    values = [_cleanup_text(v) for v in row if _cleanup_text(v)]
    rec = _empty_record()
    if not values:
        return rec

    if len(values) >= 10:
        (
            rec["name"],
            rec["customer_name"],
            rec["customer_honorific"],
            rec["customer_position"],
            rec["agency"],
            rec["branch"],
            rec["phone"],
            rec["customer_status"],
            rec["tags"],
            rec["memo2"],
        ) = values[:10]
        return rec

    if len(values) == 1:
        rec["name"] = values[0]
        return rec

    starts_with_emp_id = _looks_like_emp_id(values[0]) and len(values) >= 2 and _looks_like_nameish(values[1])
    if starts_with_emp_id:
        rec["emp_id"] = values[0]
        rec["name"] = values[1] if len(values) >= 2 else ""
        rec["phone"] = values[2] if len(values) >= 3 else ""
        rec["agency"] = values[3] if len(values) >= 4 else ""
        rec["branch"] = values[4] if len(values) >= 5 else ""
        return rec

    if len(values) == 2 and _looks_like_phone(values[1]):
        rec["name"] = values[0]
        rec["phone"] = values[1]
        return rec

    rec["name"] = values[0]
    rec["phone"] = values[1] if len(values) >= 2 else ""
    rec["agency"] = values[2] if len(values) >= 3 else ""
    rec["branch"] = values[3] if len(values) >= 4 else ""
    return rec


def _split_text_line(line: str) -> list[str]:
    raw = (line or "").rstrip("\r\n")
    if not raw.strip():
        return []

    delimiters = ("\t", "|", ";", ",")
    for delim in delimiters:
        if delim in raw:
            row = next(csv.reader(io.StringIO(raw), delimiter=delim))
            return [_cleanup_text(cell) for cell in row]

    return [_cleanup_name(raw)]


def _looks_like_phone(value: str) -> bool:
    digits = re.sub(r"\D", "", value or "")
    return len(digits) >= 8


def _looks_like_emp_id(value: str) -> bool:
    x = (value or "").strip()
    if not x or _looks_like_phone(x):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9_-]{2,20}", x))


def _looks_like_nameish(value: str) -> bool:
    x = _cleanup_text(value)
    if not x or _looks_like_phone(x):
        return False
    return bool(re.search(r"[A-Za-z가-힣]", x))


def _map_header(cell: str) -> str | None:
    key = _normalize_header(cell)
    if not key:
        return None

    for field, aliases in _HEADER_ALIASES.items():
        if key in {_normalize_header(alias) for alias in aliases}:
            return field
    return None


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s/_\-·()\[\]]+", "", (value or "").strip()).casefold()


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _cleanup_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def _cleanup_name(value: str) -> str:
    x = _cleanup_text(value)
    x = re.sub(r"^[\u2022\-\*◈◆]+\s*", "", x)
    x = re.sub(r"^\d+[\.)]\s*", "", x)
    return x.strip("\"'").strip()


def _trim_row(row: Sequence[str]) -> list[str]:
    values = [_cleanup_text(v) for v in row]
    while values and not values[-1]:
        values.pop()
    return values


def _get_cell(row: Sequence[str], idx: int | None) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return _cleanup_text(row[idx])


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag
